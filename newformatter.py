import os
import tempfile
import json
from flask import Flask, request, render_template_string, send_file, redirect, url_for, flash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = "replace_with_a_secure_key"  # Needed for flashing messages

# Temporary folder to store uploaded and processed files
UPLOAD_FOLDER = tempfile.gettempdir()
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Custom CSS for a polished look using the theme color and font.
custom_css = """
<style>
  body {
    font-family: 'Trebuchet MS', sans-serif;
    background-color: #f8f9fa;
  }
  .theme-bg {
    background-color: #1b3055;
  }
  .theme-text {
    color: #1b3055;
  }
  .btn-custom {
    background-color: #1b3055;
    border-color: #1b3055;
    color: #fff;
  }
  .btn-custom:hover {
    background-color: #153044;
    border-color: #153044;
  }
  h2 {
    color: #1b3055;
  }
  .container {
    margin-top: 50px;
    max-width: 600px;
  }
  .card {
    border: none;
    border-radius: 10px;
    box-shadow: 0 0 10px rgba(0,0,0,0.1);
  }
  .card-header {
    background-color: #1b3055;
    color: #fff;
    border-top-left-radius: 10px;
    border-top-right-radius: 10px;
  }
  .form-label {
    color: #1b3055;
  }
</style>
"""

# HTML template for file upload.
upload_template = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Excel Formatter - Upload</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  {custom_css}
</head>
<body>
  <div class="container">
    <div class="card">
      <div class="card-header text-center">
        <h2>Upload Excel File(s)</h2>
      </div>
      <div class="card-body">
        <form method="post" enctype="multipart/form-data" action="{{ url_for('upload') }}">
          <div class="mb-3">
            <input class="form-control" type="file" name="file" multiple required>
          </div>
          <div class="d-grid">
            <button type="submit" class="btn btn-custom">Upload Files</button>
          </div>
        </form>
        {% with messages = get_flashed_messages() %}
          {% if messages %}
            <div class="mt-3">
              {% for message in messages %}
                <div class="alert alert-danger">{{ message }}</div>
              {% endfor %}
            </div>
          {% endif %}
        {% endwith %}
      </div>
    </div>
  </div>
</body>
</html>
"""

# HTML template for header selection.
select_template = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Select Headers & File Name</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  {custom_css}
</head>
<body>
  <div class="container">
    <div class="card">
      <div class="card-header text-center">
        <h2>Select Headers & Enter File Name</h2>
      </div>
      <div class="card-body">
        <form method="post" action="{{ url_for('process') }}">
          <div class="mb-3">
            {% for header in headers %}
              <div class="form-check mb-2">
                <input class="form-check-input" type="checkbox" name="selected" value="{{ header }}">
                <label class="form-check-label">{{ header }}</label>
              </div>
            {% endfor %}
          </div>
          <div class="mb-3">
            <label for="desired_filename" class="form-label">Desired File Name (without extension)</label>
            <input type="text" class="form-control" id="desired_filename" name="desired_filename" placeholder="Enter file name">
          </div>
          <!-- Pass the uploaded file paths as a hidden field (JSON encoded) -->
          <input type="hidden" name="files" value='{{ files_json }}'>
          <div class="d-grid">
            <button type="submit" class="btn btn-custom">Process File</button>
          </div>
        </form>
      </div>
    </div>
  </div>
</body>
</html>
"""

def transform_text(text):
    """Convert text to proper case and perform specified replacements."""
    new_text = text.title()
    if new_text == "Bachelors Of Commerce - Commerce":
        return "B.Com. (Hons.)"
    elif new_text == "Bachelors Of Arts - Humanities":
        return "B.A. Hons. Economics"
    elif new_text == "Cgpa":
        return "CGPA"
    elif new_text == "Na":
        return "NA"
    else:
        return new_text

def process_excel(file_paths, combined_headers, desired_filename):
    """
    Process multiple Excel files:
      - Build a combined dataset using the union (in order) of headers (combined_headers)
      - For each file, ignore any "S. No." or "Serial No" columns.
      - For each file, store the actual Excel column index for each header.
      - Combine all rows from all files (filling missing values with None).
      - Transform all text (proper case, replacements) and, if "Roll No" is present, uppercase its value.
      - If "Name" is among headers, sort rows alphabetically by it.
      - Create a new workbook with a first column "S. No.", then the combined headers.
      - Add margins, a title row (merged) whose text is the desired file name,
        and apply formatting, auto-fit column widths, and hide gridlines.
    Returns the path to the processed file.
    """
    combined_data = []
    # Process each file individually.
    for path in file_paths:
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        # Build mapping: header -> actual Excel column index (1-indexed)
        file_header_map = {}
        for col_index, cell in enumerate(sheet[1], start=1):
            if cell.value is not None:
                h = cell.value
                if isinstance(h, str) and h.strip().lower() in ["s. no.", "serial no"]:
                    continue
                file_header_map[h] = col_index
        # For each data row, build a dict for combined headers.
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            row_dict = {}
            for header in combined_headers:
                if header in file_header_map:
                    excel_col = file_header_map[header]
                    row_dict[header] = row[excel_col - 1]
                else:
                    row_dict[header] = None
            combined_data.append(row_dict)
        wb.close()

    # Transform text for each cell.
    for row in combined_data:
        for header in combined_headers:
            val = row.get(header)
            if isinstance(val, str):
                row[header] = transform_text(val)
                if header.strip().lower() in ["roll no","roll number","roll no."]:
                    row[header] = row[header].upper()

    # Sort by "Name" if present.
    if any(isinstance(h, str) and h.strip().lower() == "name" for h in combined_headers):
        name_header = next(h for h in combined_headers if isinstance(h, str) and h.strip().lower() == "name")
        combined_data.sort(key=lambda r: str(r.get(name_header) or "").lower())

    # Create new workbook for output.
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active
    new_sheet.title = "Formatted Data"

    # New header order: first "S. No.", then combined_headers.
    new_order = {"S. No.": 1}
    col_num = 2
    for header in combined_headers:
        new_order[header] = col_num
        col_num += 1

    # Write header row.
    for header, new_col in new_order.items():
        new_sheet.cell(row=1, column=new_col, value=header)

    # Write data rows.
    data_rows = []
    for row_dict in combined_data:
        row_list = [None]
        for header in combined_headers:
            row_list.append(row_dict.get(header))
        data_rows.append(row_list)

    for i, row_data in enumerate(data_rows, start=2):
        for j, value in enumerate(row_data, start=1):
            new_sheet.cell(row=i, column=j, value=value)
    data_max_row = len(data_rows) + 1
    data_max_col = len(new_order)

    # Fill in Serial Numbers.
    for i in range(2, data_max_row + 1):
        new_sheet.cell(row=i, column=1, value=i - 1)

    # --- Formatting and Margins ---
    header_light_font = Font(name="Trebuchet MS", size=12, bold=True, color="FFFFFF")
    header_font = Font(name="Trebuchet MS", size=11, bold=True, color="000000")
    header_light_fill = PatternFill(start_color="1b3055", end_color="1b3055", fill_type="solid")
    header_fill = PatternFill(start_color="c9daf8", end_color="c9daf8", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    data_font = Font(name="Trebuchet MS", size=10)
    data_alignment = Alignment(horizontal="center", vertical="center")
    data_border = header_border

    # Format header row.
    for header, new_col in new_order.items():
        cell = new_sheet.cell(row=1, column=new_col)
        cell.font = header_light_font
        cell.fill = header_light_fill
        cell.alignment = header_alignment
        cell.border = header_border
        new_sheet.column_dimensions[get_column_letter(new_col)].width = 20
    new_sheet.row_dimensions[1].height = 20

    # Format data cells.
    for row in range(2, data_max_row + 1):
        for col in range(1, data_max_col + 1):
            cell = new_sheet.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = data_border

    # Insert Top Margin and Title Row.
    new_sheet.insert_rows(1, amount=2)
    new_sheet.row_dimensions[1].height = 10
    new_sheet.row_dimensions[3].height = 20
    for col in range(1, data_max_col + 1):
        cell = new_sheet.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # Insert Left, Right, and Bottom Margins.
    new_sheet.insert_cols(1)
    new_sheet.column_dimensions["A"].width = 2
    right_margin_index = data_max_col + 2
    new_sheet.insert_cols(right_margin_index)
    new_sheet.column_dimensions[get_column_letter(right_margin_index)].width = 2
    blank_row = [None] * new_sheet.max_column
    new_sheet.append(blank_row)
    new_sheet.row_dimensions[new_sheet.max_row].height = 10

    used_last_row = new_sheet.max_row
    used_last_col = new_sheet.max_column

    # Merge Title Row (Row 2) Across from Column 2 to used_last_col.
    new_sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=used_last_col)
    new_sheet.cell(row=2, column=2, value=desired_filename)

    # Delete Excess Rows and Columns.
    total_rows = new_sheet.max_row
    if total_rows > used_last_row:
        new_sheet.delete_rows(used_last_row + 1, total_rows - used_last_row)
    total_cols = new_sheet.max_column
    if total_cols > used_last_col:
        new_sheet.delete_cols(used_last_col + 1, total_cols - used_last_col)

    # Auto-fit Column Widths for Data Columns (skip margin columns).
    for col in range(2, used_last_col+1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row_cells in new_sheet.iter_rows(min_row=1, max_row=used_last_row, min_col=col, max_col=col):
            for cell in row_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
        new_sheet.column_dimensions[col_letter].width = max_length + 2

    # Hide Gridlines.
    new_sheet.sheet_view.showGridLines = False
    new_sheet.page_setup.printArea = f"A1:{get_column_letter(used_last_col)}{used_last_row}"

    out_filepath = os.path.join(app.config["UPLOAD_FOLDER"], "processed.xlsx")
    new_wb.save(out_filepath)
    new_wb.close()
    return out_filepath

@app.route("/", methods=["GET"])
def index():
    return render_template_string(upload_template.replace("{custom_css}", custom_css))

@app.route("/upload", methods=["POST"])
def upload():
    files = request.files.getlist("file")
    if not files:
        flash("No file part")
        return redirect(url_for("index"))
    combined_headers = []
    saved_paths = []
    # Process each uploaded file and build combined headers (excluding "S. No." / "Serial No")
    for file in files:
        if file.filename == "":
            continue
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
        file.save(filepath)
        saved_paths.append(filepath)
        try:
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1] if cell.value is not None]
            wb.close()
            for h in headers:
                if isinstance(h, str) and h.strip().lower() in ["s. no.", "serial no"]:
                    continue
                if h not in combined_headers:
                    combined_headers.append(h)
        except Exception as e:
            flash(f"Error reading Excel file {file.filename}: {e}")
            return redirect(url_for("index"))
    if not combined_headers:
        flash("No valid headers found in uploaded file(s).")
        return redirect(url_for("index"))
    return render_template_string(select_template.replace("{custom_css}", custom_css),
                                  headers=combined_headers,
                                  files_json=json.dumps(saved_paths),
                                  filename="")

@app.route("/process", methods=["POST"])
def process():
    selected_headers = request.form.getlist("selected")
    if not selected_headers:
        flash("Please select at least one header.")
        return redirect(url_for("index"))
    desired_filename = request.form.get("desired_filename")
    if not desired_filename:
        desired_filename = "Processed_File"
    files_json = request.form.get("files")
    try:
        file_paths = json.loads(files_json)
    except Exception as e:
        flash(f"Error processing file list: {e}")
        return redirect(url_for("index"))
    try:
        out_filepath = process_excel(file_paths, selected_headers, desired_filename)
    except Exception as e:
        flash(f"Error processing file: {e}")
        return redirect(url_for("index"))
    return send_file(out_filepath, as_attachment=True, download_name=f"{desired_filename}.xlsx")

if __name__ == "__main__":
    app.run(debug=True)
